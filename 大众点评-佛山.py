import re
import requests
import time
import openpyxl

def open_url2(url):
    headers={'User-Agent':'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/57.0.2987.98 Safari/537.36 LBBROWSER'}
    r=requests.get(url,headers=headers,timeout=5)
    html=r.text
    a=re.compile('<div class="txt".*?<h4>(.*?)</h4>.*?<div class="comment">.*?title="(.*?)".*?<b>￥(.*?)</b>.*?<span class="addr">(.*?)</span>.*?<a class="recommend-click".*?target="_blank">(.*?)</a>.*?<a class="recommend-click".*?target="_blank">(.*?)</a>.*?<a class="recommend-click".*?target="_blank">(.*?)</a>.*?<span class="comment-list">.*?<b>(.*?)</b>.*?<b>(.*?)</b>.*?<b>(.*?)</b>',re.S)
    b=re.findall(a,html)
    print(b)
    return b

def save(b):
    for each in b:
        ws.append(each)

if __name__=='__main__':
    wb = openpyxl.Workbook()
    ws = wb.active
    ws['A1'] = '店铺名'
    ws['B1'] = '店铺星级'
    ws['C1'] = '平均消费'
    ws['D1'] = '商铺地址'
    ws['E1'] = '热门菜式'
    ws['F1'] = '热门菜式'
    ws['G1'] = '热门菜式'
    ws['H1'] = '口味'
    ws['I1'] = '环境'
    ws['J1'] = '服务'
    for i in range(1,20):
        url = 'http://www.dianping.com/foshan/ch10/g103p'+str(i)
        b=open_url2(url)
        save(b)
        time.sleep(3)
        print('休息3秒后继续爬,现在爬取页数为：',str(i))
    print('爬完啦')
    wb.save('大众点评-佛山数据.xlsx')